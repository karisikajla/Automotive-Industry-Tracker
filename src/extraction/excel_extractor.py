import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..')))

import openpyxl
import chardet
from datetime import datetime
from src.utils.logger import logging
from src.storage.mongo import save_to_mongo

def detect_encoding(filepath):
    with open(filepath, "rb") as f:
        raw = f.read()
    result = chardet.detect(raw)
    encoding = result.get("encoding", "utf-8")
    logging.info(f"Detected encoding for {filepath}: {encoding}")
    return encoding

def extract_excel(filepath):
    logging.info(f"Starting Excel extraction: {filepath}")
    result = {
        "file_name": os.path.basename(filepath),
        "document_type": "excel",
        "source": filepath,
        "extraction_timestamp": datetime.utcnow().isoformat(),
        "extraction_library": "openpyxl",
        "sheets": []
    }

    try:
        detect_encoding(filepath)
        wb = openpyxl.load_workbook(filepath)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = {
                "sheet_name": sheet_name,
                "rows": []
            }

            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                sheet_data["rows"].append(row_data)

            result["sheets"].append(sheet_data)
            logging.info(f"Extracted sheet '{sheet_name}' from {os.path.basename(filepath)}")

        save_to_mongo(result, "excel_extraction")
        logging.info(f"Excel extraction complete: {filepath}")

    except Exception as e:
        logging.error(f"Error extracting Excel {filepath}: {e}")

    return result

def extract_all_excel():
    ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..'))
    excel_dir = os.path.join(ROOT_DIR, "data", "raw", "excel")
    results = []
    for filename in os.listdir(excel_dir):
        if filename.endswith(".xlsx"):
            filepath = os.path.join(excel_dir, filename)
            result = extract_excel(filepath)
            results.append(result)
    return results

results = extract_all_excel()
for r in results:
    print(f"Extracted {r['file_name']}: {len(r['sheets'])} sheets")